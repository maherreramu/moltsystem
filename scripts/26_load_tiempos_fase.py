"""
scripts/26_load_tiempos_fase.py — tiempos_por_fase.xlsx → op_ds.dias_* + recalc_pull

Lee la hoja "Producción Activa" de tiempos_por_fase.xlsx, extrae el ref de la
columna Referencia (ej. "6771-1-CAMISA MANGA LARGA" → "6771-1") y actualiza
los 8 campos dias_* en op_ds. Luego llama recalc_pull por cada OP-D actualizada
para recalcular las fechas del phase_plans.

La hoja "Pipeline Comercial" se omite (no tienen OP asignada).

Uso:
  uv run python scripts/26_load_tiempos_fase.py --dry-run
  uv run python scripts/26_load_tiempos_fase.py --apply

Flags:
  --dry-run          Muestra el plan sin escribir (default si no se pasa --apply)
  --apply            Ejecuta las actualizaciones en Supabase
  --file PATH        Ruta al Excel (default: data/tiempos_por_fase.xlsx)
  --skip-recalc      No llama recalc_pull tras actualizar (útil para debugging)
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

DEFAULT_FILE = ROOT / "data" / "tiempos_por_fase.xlsx"
SHEET_NAME = "Producción Activa"

DIAS_COLS = [
    (10, "dias_fase_0"),
    (11, "dias_compras"),
    (12, "dias_trazo"),
    (13, "dias_corte"),
    (14, "dias_tiqueteo"),
    (15, "dias_satelites"),
    (16, "dias_empaque"),
    (17, "dias_despacho"),
]


def extract_ref(referencia: str | None) -> str | None:
    """Extrae ref (op_num-seq) de "6771-1-CAMISA MANGA LARGA" → "6771-1"."""
    if not referencia:
        return None
    m = re.match(r"^(\d+)-(\d+)", str(referencia).strip())
    return f"{m.group(1)}-{m.group(2)}" if m else None


def to_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def parse_sheet(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Hoja '{SHEET_NAME}' no encontrada en {path.name}")

    ws = wb[SHEET_NAME]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ref = extract_ref(row[1])
        if not ref:
            continue
        dias = {}
        for col_idx, col_name in DIAS_COLS:
            v = to_int(row[col_idx])
            if v is not None:
                dias[col_name] = v
        if not dias:
            continue
        rows.append({"ref": ref, **dias})

    wb.close()
    return rows


def run(dry: bool, file: Path, skip_recalc: bool):
    from utils.supabase_client import get_client

    tag = "[DRY-RUN] " if dry else ""
    print("=" * 60)
    print(f"{tag}26_load_tiempos_fase — tiempos_por_fase → op_ds")
    print(f"  archivo     : {file.name}")
    print(f"  dry_run     : {dry}")
    print(f"  skip_recalc : {skip_recalc}")
    print("=" * 60)

    rows = parse_sheet(file)
    # Deduplicar: si hay varias filas con el mismo ref, quedarse con la última
    deduped: dict[str, dict] = {}
    for r in rows:
        deduped[r["ref"]] = r
    rows = list(deduped.values())
    print(f"\n  Filas parseadas (refs únicos): {len(rows)}")

    sb = get_client()
    all_refs = [r["ref"] for r in rows]
    sb_opds = (
        sb.table("op_ds")
        .select("id, ref, " + ", ".join(c for _, c in DIAS_COLS))
        .in_("ref", all_refs)
        .execute()
        .data
    )
    ref_to_opd: dict[str, dict] = {r["ref"]: r for r in sb_opds}
    not_found = [r["ref"] for r in rows if r["ref"] not in ref_to_opd]

    print(f"  Encontradas en Supabase: {len(ref_to_opd)}")
    if not_found:
        print(f"  No encontradas: {sorted(not_found)}")

    # ── Comparar y actualizar ─────────────────────────────────────────────────
    n_updated = n_identical = n_missing = 0
    n_recalc_ok = n_recalc_err = 0

    print(f"\n  {'ref':12}  {'dias_f0':>6} {'compras':>7} {'trazo':>5} {'corte':>5} {'tiqueteo':>8} {'satel':>5} {'empaque':>7} {'despacho':>8}  acción")
    print("  " + "-" * 80)

    for row in sorted(rows, key=lambda r: r["ref"]):
        ref = row["ref"]
        if ref not in ref_to_opd:
            n_missing += 1
            continue

        opd = ref_to_opd[ref]
        payload = {
            col: row[col]
            for _, col in DIAS_COLS
            if col in row and row[col] != opd.get(col)
        }

        f0 = row.get("dias_fase_0", "—")
        comp = row.get("dias_compras", "—")
        trazo = row.get("dias_trazo", "—")
        corte = row.get("dias_corte", "—")
        tiq = row.get("dias_tiqueteo", "—")
        sat = row.get("dias_satelites", "—")
        emp = row.get("dias_empaque", "—")
        des = row.get("dias_despacho", "—")

        if not payload:
            n_identical += 1
            print(f"  {ref:12}  {f0!s:>6} {comp!s:>7} {trazo!s:>5} {corte!s:>5} {tiq!s:>8} {sat!s:>5} {emp!s:>7} {des!s:>8}  igual")
            continue

        print(f"  {ref:12}  {f0!s:>6} {comp!s:>7} {trazo!s:>5} {corte!s:>5} {tiq!s:>8} {sat!s:>5} {emp!s:>7} {des!s:>8}  ✓ actualizar ({len(payload)} campos)")
        n_updated += 1

        if not dry:
            sb.table("op_ds").update(payload).eq("id", opd["id"]).execute()
            if not skip_recalc:
                try:
                    sb.rpc("recalc_pull", {"p_opd_id": opd["id"]}).execute()
                    n_recalc_ok += 1
                except Exception as e:
                    print(f"  WARN recalc_pull {ref}: {e}", flush=True)
                    n_recalc_err += 1

    print("\n" + "=" * 60)
    if dry:
        print(f"DRY-RUN: {n_updated} a actualizar, {n_identical} ya iguales, {n_missing} no existen.")
        print("Correr con --apply para ejecutar.")
    else:
        print(f"Completado: {n_updated} actualizadas, {n_identical} ya iguales, {n_missing} no existen.")
        if not skip_recalc:
            print(f"recalc_pull: {n_recalc_ok} OK, {n_recalc_err} errores.")
    print("=" * 60)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="tiempos_por_fase → op_ds.dias_*")
    parser.add_argument("--apply", action="store_true", help="Ejecutar actualizaciones (default: dry-run)")
    parser.add_argument("--dry-run", action="store_true", help="Mostrar plan sin escribir (default)")
    parser.add_argument("--file", default=str(DEFAULT_FILE), help="Ruta al Excel")
    parser.add_argument("--skip-recalc", action="store_true", help="No llamar recalc_pull")
    args = parser.parse_args()
    run(dry=not args.apply, file=Path(args.file), skip_recalc=args.skip_recalc)
