"""
scripts/25_load_vista_maestra.py — Vista Maestra → ops.fecha_compromiso + ops.comercial

Lee la hoja "Vista Maestra" del archivo MOLT_Vista_Integrada_Proyectos.xlsx
y actualiza fecha_compromiso (y opcionalmente comercial) en la tabla ops.

Regla: solo actualiza si la celda tiene valor; nunca pisa con None.
Celdas de "No. OP" pueden contener múltiples OPs ("6797, 6787 y 6786") —
se extraen todos los números de 4+ dígitos.

Uso:
  uv run python scripts/25_load_vista_maestra.py --dry-run
  uv run python scripts/25_load_vista_maestra.py --apply

Flags:
  --dry-run        Muestra el plan sin escribir (default si no se pasa --apply)
  --apply          Ejecuta las actualizaciones en Supabase
  --file PATH      Ruta al Excel (default: data/MOLT_Vista_Integrada_Proyectos.xlsx)
  --skip-comercial No actualiza el campo comercial (solo fecha_compromiso)
"""

import argparse
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

DEFAULT_FILE = ROOT / "data" / "MOLT_Vista_Integrada_Proyectos.xlsx"
SHEET_NAME = "Vista Maestra"
HEADER_ROW = 3
DATA_START = 4


def extract_op_nums(val) -> list[str]:
    """Extrae todos los op_nums (4+ dígitos) de un valor de celda."""
    if val is None:
        return []
    return re.findall(r"\d{4,}", str(val))


def to_iso_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date().isoformat() if isinstance(val, datetime) else val.isoformat()
    try:
        return date.fromisoformat(str(val)[:10]).isoformat()
    except (ValueError, TypeError):
        return None


def parse_vista_maestra(path: Path) -> list[dict]:
    """Lee Vista Maestra y retorna lista de {op_nums, fecha_compromiso, comercial}."""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Hoja '{SHEET_NAME}' no encontrada en {path.name}")

    ws = wb[SHEET_NAME]
    rows = []
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        op_raw = row[2]   # No. OP
        comercial = row[3]  # Responsable
        fecha_comp = row[5]  # Fecha Compromiso

        op_nums = extract_op_nums(op_raw)
        if not op_nums:
            continue

        rows.append({
            "op_nums":          op_nums,
            "op_raw":           str(op_raw),
            "fecha_compromiso": to_iso_date(fecha_comp),
            "comercial":        str(comercial).strip() if comercial else None,
        })

    wb.close()
    return rows


def run(dry: bool, file: Path, skip_comercial: bool):
    from utils.supabase_client import get_client

    tag = "[DRY-RUN] " if dry else ""
    print("=" * 60)
    print(f"{tag}25_load_vista_maestra — Vista Maestra → ops")
    print(f"  archivo : {file.name}")
    print(f"  dry_run : {dry}")
    print("=" * 60)

    rows = parse_vista_maestra(file)
    print(f"\n  Filas parseadas: {len(rows)}")
    total_ops = sum(len(r["op_nums"]) for r in rows)
    print(f"  op_nums totales: {total_ops}")

    sb = get_client()
    all_op_nums = [n for r in rows for n in r["op_nums"]]
    sb_ops = sb.table("ops").select("op_num, fecha_compromiso, comercial").in_("op_num", all_op_nums).execute().data
    found_set = {r["op_num"] for r in sb_ops}
    not_found = [n for n in all_op_nums if n not in found_set]

    print(f"  Encontradas en Supabase: {len(found_set)}")
    if not_found:
        print(f"  No encontradas (no están en Supabase aún): {sorted(not_found)}")

    print(f"\n{'Op':8}  {'fecha_comp_actual':18}  {'fecha_comp_nueva':18}  {'comercial':25}  acción")
    print("-" * 90)

    n_updated = n_skipped = n_missing = 0
    for row in rows:
        for op_num in row["op_nums"]:
            if op_num not in found_set:
                n_missing += 1
                print(f"  {op_num:6}  {'—':18}  {row['fecha_compromiso'] or '—':18}  {'—':25}  NO EXISTE")
                continue

            db_row = next(r for r in sb_ops if r["op_num"] == op_num)
            payload: dict = {}

            if row["fecha_compromiso"]:
                payload["fecha_compromiso"] = row["fecha_compromiso"]

            if not skip_comercial and row["comercial"]:
                payload["comercial"] = row["comercial"]

            if not payload:
                n_skipped += 1
                print(f"  {op_num:6}  {str(db_row['fecha_compromiso'] or '—'):18}  {'—':18}  {'—':25}  sin datos")
                continue

            fc_actual = str(db_row["fecha_compromiso"] or "—")
            fc_nueva = row["fecha_compromiso"] or "—"
            comercial_str = (row["comercial"] or "—") if not skip_comercial else "(omitido)"
            print(f"  {op_num:6}  {fc_actual:18}  {fc_nueva:18}  {comercial_str:25}  ✓ actualizar")

            if not dry:
                sb.table("ops").update(payload).eq("op_num", op_num).execute()
            n_updated += 1

    print("\n" + "=" * 60)
    if dry:
        print(f"DRY-RUN: {n_updated} ops a actualizar, {n_skipped} sin datos, {n_missing} no existen.")
        print("Correr con --apply para ejecutar.")
    else:
        print(f"Completado: {n_updated} ops actualizadas, {n_skipped} sin datos, {n_missing} no existen.")
    print("=" * 60)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Vista Maestra → ops.fecha_compromiso")
    parser.add_argument("--apply", action="store_true", help="Ejecutar actualizaciones (default: dry-run)")
    parser.add_argument("--dry-run", action="store_true", help="Mostrar plan sin escribir (default)")
    parser.add_argument("--file", default=str(DEFAULT_FILE), help="Ruta al Excel")
    parser.add_argument("--skip-comercial", action="store_true", help="No actualizar campo comercial")
    args = parser.parse_args()
    run(dry=not args.apply, file=Path(args.file), skip_comercial=args.skip_comercial)
